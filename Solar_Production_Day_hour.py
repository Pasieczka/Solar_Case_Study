#This script will find the day with the worst solar production
#We want this value so we are able to determine how to size our grid accrodingly
#This script takes a CSV file and manipulates the data to view it easier

#moving data to get Worst Day solar production

import datetime
import openpyxl
import csv
import sys
import pandas as pd

print("the start")

wb = openpyxl.Workbook()
sheet = wb.active


#open solar data and read information
f = open("Solar_Production_Data_Year.csv", 'r')
reader = csv.reader(f)
kw_value = []
for row in reader:
    kw_value.append(row[1])
#print(kw_value)

#Definign Row Titles

time_hours = datetime.datetime(2014,1,1,0,0,0) #start date
time_interval = datetime.timedelta(minutes=5)



sheet.cell(row = 1, column = 1).value = 'Time'

#Updated way of adding time to column 1 in excel
hour = 0
for i in range (2,26):
    sheet.cell(row = i, column = 1).value = hour #adding time interval to spreadsheet
    hour += 1

#ADDING DATES
days = pd.date_range(start="2014-01-01", end="2014-12-31").to_pydatetime().tolist()
day_column = 2
for day in days:
    sheet.cell(row = 1, column = day_column).value = day.strftime("%m-%d")
    day_column += 1
    
    
column_number = 2
time_counter = 2
hour_instant = 0
hour_tracker = 0
day_sum = 0
lowest_production_value = 100000
lowest_production_day = "none"
    

#MOVING CSV SOLAR DATA TO BETTER FORMAT
with open("Solar_Production_Data_Year.csv", 'r') as file:
    
    for line in file:
        line = line.split(',')
        hour_instant += float(line[1])
        hour_tracker +=1

        if hour_tracker == 12:
            
            hour_instant_average = (hour_instant/12) #getting kWh
            sheet.cell(row = time_counter, column = column_number).value = hour_instant_average
            
            day_sum += hour_instant_average #add the hour value to the running sume
            hour_tracker = 0 #reset hour value 
            hour_instant = 0 #reset hour averge
            time_counter +=1

            

            if time_counter == 26: #number of time intervals that day
                #put sum at bottom of row
                daily_kw_average = day_sum/24 #get average kWh for that day. 
                sheet.cell(row = 30, column = column_number).value = (daily_kw_average) 
                
                #reset counter values
                column_number +=1
                time_counter = 2
                day_sum = 0
                
                if daily_kw_average <= lowest_production_value:
                    print("in if statement")
                    print("new lowest value is: " + str(daily_kw_average))
                    lowest_production_day = str(line[0])
                    lowest_production_value = daily_kw_average 
print("")
print("Day with lowest production day is: " + str(lowest_production_day[:9]))#str(sheet[column_number])) 
print("Lowest production value in kW is: " + str(lowest_production_value))       
spreadsheet = wb.save('Worst_Solar_Day_Hourly.csv')
           
print('End of script')
