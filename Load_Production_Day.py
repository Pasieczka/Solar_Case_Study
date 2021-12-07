#THIS SCRIPT IS USED TO FIND DAILY LOAD IN kW AND kVA. 
#Created by Nicholas Pasieczka
#**********************************

import datetime
import openpyxl
import csv
import sys
import pandas as pd

print("the start")

wb = openpyxl.Workbook()
sheet = wb.active


#open load data and read information
f = open("Load_Data_Day_Clean.csv", 'r')
reader = csv.reader(f)
day_list = []
kw_value = []
kva_value = []

#moving data values to a list
for row in reader:
    day_list.append(row[0])
    kw_value.append(row[1])
    kva_value.append(row[2])
print("Maximum kW load is: " + str(min(kw_value)))
print("Maximum kw load occurs on line : " + str(kw_value.index(min(kw_value))))

print("Maximum kVA load is: " + str(min(kva_value)))
print("Maximum kVA load occurs on line : " + str(kva_value.index(min(kva_value))))
# print("Maximum kVA load is: " + str(max(kva_value)))


#Setting up spreadsheet   
sheet.cell(row = 1, column = 1).value = 'Time'
sheet.cell(row = 35, column = 1).value = 'Time'
sheet.cell(row = 29, column = 1).value = 'Average hourly kW load'
sheet.cell(row = 30, column = 1).value = 'Daily Total kW load'
sheet.cell(row = 60, column = 1).value = 'Average hourly kVA load'
sheet.cell(row = 61, column = 1).value = 'Daily Total kVA load'
sheet.column_dimensions['A'].width = 20

hour = 0
#Updated way of adding time to excel
for i in range (2,26):
    sheet.cell(row = i, column = 1).value = hour #adding time interval to spreadsheet
    hour += 1
    
hour = 0  
#PUT THESE INTO FUNCTIONS LATER!!!!!!!
for i in range (34,58):
    sheet.cell(row = i, column = 1).value = hour #adding time interval to spreadsheet
    hour += 1
    
    
#ADDING DATES
days = pd.date_range(start="1990-01-01", end="1990-12-31").to_pydatetime().tolist()
day_column = 2
for day in days:
    sheet.cell(row = 1, column = day_column).value = day.strftime("%m-%d")
    day_column += 1
    

#Setting counter values
column_number = 2
time_counter = 2
day_kw_total = 0
day_kva_total = 0
highest_kw_production_value = 1
highest_kva_production_value = 1
highest_kw_production_day = "none"
highest_kva_production_day = "none"


#MOVING CSV LOAD DATA TO AN MORE MANIPULABLE FORMAT
with open("Load_Data_Day.csv", 'r') as file:
    
    for line in file:
        line = line.split(',')
        #print(line)
        kva_line = time_counter + 32
        sheet.cell(row = time_counter, column = column_number).value = float(line[1])  #kW value
        sheet.cell(row = kva_line, column = column_number).value = float(line[2])  #kVA value
        time_counter +=1
        day_kw_total += float(line[1])
        day_kva_total += float(line[2])
        
        if time_counter == 26: #number of time intervals that day
           
            #put sum at bottom of row in excel
            daily_kw_average = day_kw_total/24 #get average kWh for that day.
            daily_kva_average = day_kva_total/24 #get average kVA for that day.
            sheet.cell(row = 29, column = column_number).value = daily_kw_average 
            sheet.cell(row = 30, column = column_number).value = day_kw_total 
            sheet.cell(row = 60, column = column_number).value = daily_kva_average
            sheet.cell(row = 61, column = column_number).value = day_kva_total
            
            #reset counter values
                
            
            #Determine highest kW value and day it occured on
            if daily_kw_average >= highest_kw_production_value:
                # print("in highest kw statement")
                # print("new highest kw value is: " + str(daily_kw_average))
                highest_kw_production_day = str(line[0])
                highest_kw_production_value = daily_kw_average 
                
            #Determine highest value of KVA and date it occured on
            if daily_kva_average >= highest_kva_production_value:
                # print("in highest kva statement")
                # print("new highest kva value is: " + str(daily_kva_average))
                highest_kva_production_day = str(line[0])
                highest_kva_production_value = daily_kva_average 
            
            column_number +=1
            time_counter = 2
            day_kva_total = 0
            day_kw_total = 0

print("")
              
#Printing values at the end
print("Day with highest kW average load day is: " + str(highest_kw_production_day[:9]))
print("Highest daily average load requirements in kW is: " + str('{0:.7g}'.format(highest_kw_production_value))) 
print("")

print("Day with highest average kVA load day is: " + str(highest_kva_production_day[:9]))
print("Highest daily average load requirements in kVA is: " + str('{0:.7g}'.format(highest_kva_production_value))) 
      

spreadsheet = wb.save('Daily_Load.csv')
    
     

print('End of script')
